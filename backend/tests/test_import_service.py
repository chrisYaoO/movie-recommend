from datetime import date
import unittest
from tempfile import TemporaryDirectory
from pathlib import Path

from openpyxl import Workbook

from backend.app.services.import_service import (
    EXPECTED_COLUMNS,
    InMemoryViewingHistoryRawRepository,
    ViewingHistoryImportService,
    map_raw_viewing_history,
    read_viewing_history_excel,
    read_viewing_history_excel_with_preview,
    stable_row_hash,
)


class ViewingHistoryImportServiceTest(unittest.TestCase):
    def setUp(self) -> None:
        self.repository = InMemoryViewingHistoryRawRepository()
        self.service = ViewingHistoryImportService(self.repository)

    def test_import_preserves_raw_excel_shape_and_generates_hash(self) -> None:
        result = self.service.import_rows(
            "history.xlsx",
            [
                {
                    "Date": "2026-05-12",
                    "Name": "  Yi Yi  ",
                    "Director": "Edward Yang",
                    "Year": 2000,
                    "Rating": 5.0,
                    "Quality": "1080p",
                    "Comment": " favorite ",
                    "DoubanSubjectId": 1291561,
                    "DoubanImageId": 123456,
                }
            ],
        )

        self.assertEqual(1, result.imported_count)
        row = result.rows[0]
        self.assertEqual("history.xlsx", row.source_file)
        self.assertEqual(1, row.source_row_number)
        self.assertEqual("Yi Yi", row.name_raw)
        self.assertEqual("2000", row.year_raw)
        self.assertEqual("5.0", row.rating_raw)
        self.assertEqual("favorite", row.comment_raw)
        self.assertEqual("1291561", row.douban_subject_id_raw)
        self.assertEqual("123456", row.douban_image_id_raw)
        self.assertEqual(64, len(row.source_row_hash))

    def test_repeated_import_skips_duplicate_rows_by_stable_hash(self) -> None:
        rows = [
            {
                "Date": "2026-05-12",
                "Name": "Yi Yi",
                "Director": "Edward Yang",
                "Year": "2000",
                "Rating": "5.0",
                "Quality": "1080p",
                "Comment": "favorite",
            }
        ]

        first = self.service.import_rows("history.xlsx", rows)
        second = self.service.import_rows("history.xlsx", rows)

        self.assertEqual(1, first.imported_count)
        self.assertEqual(0, second.imported_count)
        self.assertEqual(1, second.skipped_duplicate_count)
        self.assertEqual(1, len(self.repository.all()))

    def test_import_rows_requires_name_and_numeric_rating(self) -> None:
        result = self.service.import_rows(
            "history.xlsx",
            [
                {"Name": "Yi Yi", "Rating": 5.0},
                {"Name": "No Rating"},
                {"Rating": 4.0},
                {"Name": "Unknown Rating", "Rating": "/"},
            ],
        )

        self.assertEqual(1, result.imported_count)
        self.assertEqual(3, result.skipped_invalid_count)
        self.assertEqual("Yi Yi", result.rows[0].name_raw)

    def test_hash_is_independent_of_extra_columns_and_whitespace(self) -> None:
        left = {column: "" for column in EXPECTED_COLUMNS}
        left.update({"Name": "Yi Yi", "Director": "Edward Yang"})
        right = {column: None for column in EXPECTED_COLUMNS}
        right.update({"Name": " Yi   Yi ", "Director": " Edward Yang ", "Ignored": "value"})

        left_hash = stable_row_hash({column: left.get(column) for column in EXPECTED_COLUMNS})
        normalized_right = self.service._to_raw_row("history.xlsx", 1, right)

        self.assertEqual(left_hash, normalized_right.source_row_hash)

    def test_read_excel_reads_all_sheets_maps_columns_and_filters_invalid_rows(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "history.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "2026"
            sheet.append(["Ignored", "Date", "Name", "Director", "Year", "Rating", "Quality", "Comment", "movie_id", "image_id"])
            sheet.append(["x", "2026-05-12", "Yi Yi", "Edward Yang", 2000, 5.0, "1080p", "favorite", 1291561, 456789])
            sheet.append(["x", "2026-05-13", "No Rating", "Nobody", 2001, None, "1080p", "skip"])
            sheet.append(["x", "2026-05-14", "Unknown Rating", "Nobody", 2002, "/", "1080p", "skip"])
            sheet.append([None, None, None, None, None, None, None, None])

            older_sheet = workbook.create_sheet("2025")
            older_sheet.append(["Date", "Name", "Director", "Year", "Ratings", "Quality", "Comments"])
            older_sheet.append(["2025-01-01", "Still Walking", "Hirokazu Kore-eda", 2008, 4.5, None, "good"])
            older_sheet.append(["2025-01-02", "Only Name", None, None, None, None, None])

            legacy_sheet = workbook.create_sheet("Sheet1")
            legacy_sheet.append(["2024-11-01", "Old Format", "Director", 2023, 3.8])

            workbook.save(path)

            rows = read_viewing_history_excel(path)
            read_result = read_viewing_history_excel_with_preview(path)
            result = self.service.import_excel(path)

        self.assertEqual(3, len(rows))
        self.assertEqual(3, read_result.preview.valid_count)
        self.assertEqual(2, read_result.preview.invalid_counts["missing_rating"])
        self.assertEqual(1, read_result.preview.invalid_counts["non_numeric_rating"])
        self.assertEqual(0, read_result.preview.blank_count)
        self.assertEqual(1, read_result.preview.sheets[2].valid_count)
        self.assertEqual("Yi Yi", rows[0]["Name"])
        self.assertEqual("Still Walking", rows[1]["Name"])
        self.assertEqual("4.5", result.rows[1].rating_raw)
        self.assertEqual("Old Format", rows[2]["Name"])
        self.assertEqual(3, result.imported_count)
        self.assertEqual(0, result.skipped_invalid_count)
        self.assertEqual("history.xlsx#2026", result.rows[0].source_file)
        self.assertEqual(2, result.rows[0].source_row_number)
        self.assertEqual("1291561", result.rows[0].douban_subject_id_raw)
        self.assertEqual("456789", result.rows[0].douban_image_id_raw)
        self.assertEqual("history.xlsx#Sheet1", result.rows[2].source_file)

    def test_preview_excel_does_not_import_rows(self) -> None:
        with TemporaryDirectory() as directory:
            path = Path(directory) / "history.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Date", "Name", "Director", "Year", "Rating", "Quality", "Comment"])
            sheet.append(["2026-05-12", "Yi Yi", "Edward Yang", 2000, 5.0, "1080p", "favorite"])
            workbook.save(path)

            preview = self.service.preview_excel(path)

        self.assertEqual(1, preview.valid_count)
        self.assertEqual(0, len(self.repository.all()))

    def test_maps_raw_rows_to_viewing_history_candidates(self) -> None:
        result = self.service.import_rows(
            "history.xlsx",
            [
                {
                    "Date": "2026-05-12 00:00:00",
                    "Name": "  Yi Yi  ",
                    "Director": "Edward Yang",
                    "Year": 2000.0,
                    "Rating": "5.0",
                    "Quality": "1080p",
                    "Comment": " favorite ",
                    "DoubanSubjectId": 1291561.0,
                    "DoubanImageId": 456789.0,
                }
            ],
        )

        mapping = map_raw_viewing_history(result.rows)

        self.assertEqual(1, len(mapping.candidates))
        self.assertEqual([], mapping.issues)
        candidate = mapping.candidates[0]
        self.assertEqual(result.rows[0].id, candidate.source_raw_id)
        self.assertEqual(result.rows[0].source_row_hash, candidate.source_row_hash)
        self.assertEqual("Yi Yi", candidate.title)
        self.assertEqual(date(2026, 5, 12), candidate.watched_date)
        self.assertEqual(2000, candidate.release_year)
        self.assertEqual(5.0, candidate.user_rating)
        self.assertEqual("Edward Yang", candidate.director)
        self.assertEqual("favorite", candidate.comment)
        self.assertEqual("1291561", candidate.douban_subject_id)
        self.assertEqual("456789", candidate.douban_image_id)

    def test_service_maps_repository_rows_to_candidates(self) -> None:
        self.service.import_rows(
            "history.xlsx",
            [{"Name": "Yi Yi", "Rating": 5.0}],
        )

        mapping = self.service.to_viewing_history_candidates()

        self.assertEqual(1, len(mapping.candidates))
        self.assertEqual("Yi Yi", mapping.candidates[0].title)


if __name__ == "__main__":
    unittest.main()
