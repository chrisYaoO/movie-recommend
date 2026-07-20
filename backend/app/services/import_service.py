from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from itertools import chain
from pathlib import Path
from typing import Any

from backend.app.models.domain import ViewingHistoryCandidate, ViewingHistoryRaw

EXPECTED_COLUMNS = ("Date", "Name", "Director", "Year", "Rating", "Quality", "Comment")
EXTERNAL_ID_COLUMNS = ("DoubanSubjectId", "DoubanImageId")
RAW_HASH_COLUMNS = (*EXPECTED_COLUMNS, *EXTERNAL_ID_COLUMNS)
COLUMN_ALIASES = {
    "Ratings": "Rating",
    "Comments": "Comment",
    "movie_id": "DoubanSubjectId",
    "image_id": "DoubanImageId",
}
SOURCE_SHEET_KEY = "__source_sheet"
SOURCE_ROW_NUMBER_KEY = "__source_row_number"


@dataclass(frozen=True)
class ImportResult:
    imported_count: int
    skipped_duplicate_count: int
    skipped_invalid_count: int
    rows: list[ViewingHistoryRaw]


@dataclass(frozen=True)
class SheetImportSummary:
    sheet_name: str
    valid_count: int
    blank_count: int
    invalid_counts: dict[str, int]

    @property
    def invalid_count(self) -> int:
        return sum(self.invalid_counts.values())

    @property
    def scanned_count(self) -> int:
        return self.valid_count + self.blank_count + self.invalid_count


@dataclass(frozen=True)
class ImportPreview:
    source_file: str
    valid_count: int
    blank_count: int
    invalid_counts: dict[str, int]
    sheets: list[SheetImportSummary]

    @property
    def invalid_count(self) -> int:
        return sum(self.invalid_counts.values())


@dataclass(frozen=True)
class ExcelReadResult:
    rows: list[dict[str, Any]]
    preview: ImportPreview


@dataclass(frozen=True)
class ViewingHistoryMappingIssue:
    source_raw_id: str
    source_sheet_name: str
    source_row_number: int
    reason: str


@dataclass(frozen=True)
class ViewingHistoryMappingResult:
    candidates: list[ViewingHistoryCandidate]
    issues: list[ViewingHistoryMappingIssue]


class InMemoryViewingHistoryRawRepository:
    def __init__(self) -> None:
        self.rows_by_source: dict[tuple[str, int], ViewingHistoryRaw] = {}

    def add_if_absent(self, row: ViewingHistoryRaw) -> bool:
        key = (row.source_sheet_name, row.source_row_number)
        if key in self.rows_by_source:
            return False
        self.rows_by_source[key] = row
        return True

    def all(self) -> list[ViewingHistoryRaw]:
        return list(self.rows_by_source.values())


class ViewingHistoryImportService:
    def __init__(self, repository: InMemoryViewingHistoryRawRepository) -> None:
        self.repository = repository

    def import_rows(self, source_name: str, rows: list[dict[str, Any]]) -> ImportResult:
        imported: list[ViewingHistoryRaw] = []
        skipped_duplicates = 0
        skipped_invalid = 0

        for index, row in enumerate(rows, start=1):
            if not _has_required_viewing_history_fields(row):
                skipped_invalid += 1
                continue

            source_sheet = _normalize_cell(row.get(SOURCE_SHEET_KEY))
            row_number = row.get(SOURCE_ROW_NUMBER_KEY, index)
            source_sheet_name = source_sheet or source_name
            raw = self._to_raw_row(source_sheet_name, int(row_number), row)
            if self.repository.add_if_absent(raw):
                imported.append(raw)
            else:
                skipped_duplicates += 1

        return ImportResult(
            imported_count=len(imported),
            skipped_duplicate_count=skipped_duplicates,
            skipped_invalid_count=skipped_invalid,
            rows=imported,
        )

    def import_excel(self, path: str | Path) -> ImportResult:
        excel_path = Path(path)
        rows = read_viewing_history_excel(excel_path)
        return self.import_rows(excel_path.name, rows)

    def preview_excel(self, path: str | Path) -> ImportPreview:
        excel_path = Path(path)
        return read_viewing_history_excel_with_preview(excel_path).preview

    def to_viewing_history_candidates(self) -> ViewingHistoryMappingResult:
        return map_raw_viewing_history(self.repository.all())

    def _to_raw_row(self, source_sheet_name: str, row_number: int, row: dict[str, Any]) -> ViewingHistoryRaw:
        values = {column: _normalize_cell(row.get(column)) for column in RAW_HASH_COLUMNS}
        return ViewingHistoryRaw(
            source_sheet_name=source_sheet_name,
            source_row_number=row_number,
            source_row_checksum=stable_row_hash(values),
            date_raw=values["Date"],
            name_raw=values["Name"],
            director_raw=values["Director"],
            year_raw=values["Year"],
            rating_raw=values["Rating"],
            quality_raw=values["Quality"],
            comment_raw=values["Comment"],
            douban_subject_id_raw=values["DoubanSubjectId"],
            douban_image_id_raw=values["DoubanImageId"],
        )


def stable_row_hash(values: dict[str, str | None]) -> str:
    canonical = "\n".join(f"{column}={values.get(column) or ''}" for column in RAW_HASH_COLUMNS)
    return sha256(canonical.encode("utf-8")).hexdigest()


def map_raw_viewing_history(rows: list[ViewingHistoryRaw]) -> ViewingHistoryMappingResult:
    candidates: list[ViewingHistoryCandidate] = []
    issues: list[ViewingHistoryMappingIssue] = []

    for row in rows:
        title = _normalize_cell(row.name_raw)
        rating = _parse_float(row.rating_raw)
        if title is None:
            issues.append(_mapping_issue(row, "missing_name"))
            continue
        if rating is None:
            issues.append(_mapping_issue(row, "invalid_rating"))
            continue

        candidates.append(
            ViewingHistoryCandidate(
                source_raw_id=row.id,
                source_sheet_name=row.source_sheet_name,
                source_row_number=row.source_row_number,
                title=title,
                user_rating=rating,
                source_row_checksum=row.source_row_checksum,
                watched_date=_parse_date(row.date_raw, row.source_sheet_name),
                director=_normalize_cell(row.director_raw),
                release_year=_parse_int(row.year_raw),
                quality=_normalize_cell(row.quality_raw),
                comment=_normalize_cell(row.comment_raw),
                douban_subject_id=_normalize_external_id(row.douban_subject_id_raw),
                douban_image_id=_normalize_external_id(row.douban_image_id_raw),
            )
        )

    return ViewingHistoryMappingResult(candidates=candidates, issues=issues)


def _mapping_issue(row: ViewingHistoryRaw, reason: str) -> ViewingHistoryMappingIssue:
    return ViewingHistoryMappingIssue(
        source_raw_id=row.id,
        source_sheet_name=row.source_sheet_name,
        source_row_number=row.source_row_number,
        reason=reason,
    )


def read_viewing_history_excel(path: str | Path) -> list[dict[str, Any]]:
    return read_viewing_history_excel_with_preview(path).rows


def read_viewing_history_excel_with_preview(path: str | Path) -> ExcelReadResult:
    from openpyxl import load_workbook

    excel_path = Path(path)
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    sheet_summaries: list[SheetImportSummary] = []

    try:
        for sheet in workbook.worksheets:
            sheet_result = _read_sheet_rows(sheet)
            rows.extend(sheet_result.rows)
            sheet_summaries.append(sheet_result.summary)
    finally:
        workbook.close()

    invalid_counts: dict[str, int] = {}
    for summary in sheet_summaries:
        for reason, count in summary.invalid_counts.items():
            invalid_counts[reason] = invalid_counts.get(reason, 0) + count

    preview = ImportPreview(
        source_file=excel_path.name,
        valid_count=len(rows),
        blank_count=sum(summary.blank_count for summary in sheet_summaries),
        invalid_counts=invalid_counts,
        sheets=sheet_summaries,
    )
    return ExcelReadResult(rows=rows, preview=preview)


@dataclass(frozen=True)
class SheetReadResult:
    rows: list[dict[str, Any]]
    summary: SheetImportSummary


def _read_sheet_rows(sheet: Any) -> SheetReadResult:
    iterator = sheet.iter_rows(values_only=True)
    try:
        first_row = next(iterator)
    except StopIteration:
        return SheetReadResult(
            rows=[],
            summary=SheetImportSummary(sheet_name=sheet.title, valid_count=0, blank_count=0, invalid_counts={}),
        )

    first_values = list(first_row)
    header = [_normalize_cell(value) for value in first_values]
    has_header = "Name" in [_canonical_column_name(value) for value in header]
    column_indexes = _resolve_column_indexes(header) if has_header else _legacy_column_indexes()
    data_iterator = iterator if has_header else chain([tuple(first_values)], iterator)

    rows: list[dict[str, Any]] = []
    blank_count = 0
    pending_blank_count = 0
    invalid_counts: dict[str, int] = {}
    for row_number, excel_row in enumerate(data_iterator, start=2 if has_header else 1):
        row_values = list(excel_row)
        if not any(_normalize_cell(value) is not None for value in row_values):
            pending_blank_count += 1
            continue

        blank_count += pending_blank_count
        pending_blank_count = 0
        row = {
            column: row_values[index] if index < len(row_values) else None
            for column, index in column_indexes.items()
        }
        row[SOURCE_SHEET_KEY] = sheet.title
        row[SOURCE_ROW_NUMBER_KEY] = row_number
        invalid_reason = _invalid_viewing_history_reason(row)
        if invalid_reason is None:
            rows.append(row)
        else:
            invalid_counts[invalid_reason] = invalid_counts.get(invalid_reason, 0) + 1

    return SheetReadResult(
        rows=rows,
        summary=SheetImportSummary(
            sheet_name=sheet.title,
            valid_count=len(rows),
            blank_count=blank_count,
            invalid_counts=invalid_counts,
        ),
    )


def _resolve_column_indexes(header: list[str | None]) -> dict[str, int]:
    indexes = {}
    for index, name in enumerate(header):
        canonical = _canonical_column_name(name)
        if canonical in RAW_HASH_COLUMNS and canonical not in indexes:
            indexes[canonical] = index

    if "Name" not in indexes or "Rating" not in indexes:
        raise ValueError("Missing required Excel columns: Name, Rating")

    return {
        column: indexes[column]
        for column in RAW_HASH_COLUMNS
        if column in indexes
    }


def _legacy_column_indexes() -> dict[str, int]:
    return {column: index for index, column in enumerate(EXPECTED_COLUMNS[:5])}


def _canonical_column_name(value: str | None) -> str | None:
    if value is None:
        return None
    return COLUMN_ALIASES.get(value, value)


def _has_required_viewing_history_fields(row: dict[str, Any]) -> bool:
    return _invalid_viewing_history_reason(row) is None


def _invalid_viewing_history_reason(row: dict[str, Any]) -> str | None:
    if _normalize_cell(row.get("Name")) is None:
        return "missing_name"
    if _normalize_cell(row.get("Rating")) is None:
        return "missing_rating"
    if not _is_numeric_rating(row.get("Rating")):
        return "non_numeric_rating"
    return None


def _is_numeric_rating(value: Any) -> bool:
    normalized = _normalize_cell(value)
    if normalized is None:
        return False
    return _parse_float(normalized) is not None


def _parse_float(value: Any) -> float | None:
    normalized = _normalize_cell(value)
    if normalized is None:
        return None
    try:
        return float(normalized)
    except ValueError:
        return None


def _parse_int(value: Any) -> int | None:
    normalized = _normalize_cell(value)
    if normalized is None:
        return None
    try:
        return int(float(normalized))
    except ValueError:
        return None


def _parse_date(value: Any, source_sheet_name: str | None = None) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    normalized = _normalize_cell(value)
    if normalized is None:
        return None

    for format_text in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(normalized, format_text).date()
        except ValueError:
            continue
    if source_sheet_name and len(source_sheet_name) == 4 and source_sheet_name.isdigit():
        try:
            return datetime.strptime(f"{source_sheet_name}/{normalized}", "%Y/%m/%d").date()
        except ValueError:
            pass
    return None


def _normalize_external_id(value: Any) -> str | None:
    parsed = _parse_int(value)
    if parsed is not None:
        return str(parsed)
    return _normalize_cell(value)


def _normalize_cell(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return " ".join(text.split())
